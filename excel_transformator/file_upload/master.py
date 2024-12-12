from openpyxl.reader.excel import load_workbook

from .models import Categories, Companies, Lob


def process_excel_file(file_path):
    # Load the Excel file
    workbook = load_workbook(file_path)
    lob_list = []
    row_data = dict()
    # Iterate through all sheets
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if sheet_name in ['month']:
                if 'month' not in row_data.keys():
                    pass
            if sheet_name in ['category']:
                Categories.objects.create(clubbed_name=row[0].strip(), category=row[1].strip())
            elif sheet_name == 'name':
                Companies.objects.create(insurer=row[0].strip(), name=row[1].strip(), clubbed_name=row[2].strip())
            elif sheet_name == 'lob':
                lob_list.append(row[0].strip())
    row_data['lob_list'] = lob_list
    Lob.objects.create(lob=lob_list)

#   file path --> "C:/Users/prasa/Downloads/master.xlsx"